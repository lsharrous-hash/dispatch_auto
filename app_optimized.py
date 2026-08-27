import streamlit as st
import pandas as pd
import folium
from folium.plugins import Draw, FastMarkerCluster
from streamlit_folium import st_folium
import json
import os
from datetime import datetime
from data_processor import load_data, preparer_telechargement_excel
from shapely.geometry import shape, Point
import zipfile
import io
import unicodedata
import re

# Configuration
st.set_page_config(layout="wide", page_title="Dispatch Auto - JNR Transport")

PATTERNS_FILE = "driver_patterns.json"

# === CACHE ET OPTIMISATIONS ===

@st.cache_data
def load_and_process_file(file_content, file_name):
    """Cache le chargement des fichiers."""
    import io as io_module
    
    # Forcer la détection par extension
    file_ext = file_name.lower().split('.')[-1] if '.' in file_name else ''
    
    if file_ext in ['xlsx', 'xls']:
        # Essayer plusieurs méthodes de lecture pour les fichiers Excel problématiques
        df = None
        
        # Méthode 1: openpyxl avec data_only=True (ignore les formules, lit les valeurs)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io_module.BytesIO(file_content), data_only=True, read_only=True)
            ws = wb.active
            
            # Lire les données manuellement
            data = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    # Vérifier si c'est une ligne d'en-tête valide ou une ligne vide
                    if row[0] is None or str(row[0]).startswith('Unnamed'):
                        continue
                    headers = [str(c) if c else f'Col_{j}' for j, c in enumerate(row)]
                else:
                    if headers is None:
                        headers = [str(c) if c else f'Col_{j}' for j, c in enumerate(row)]
                    else:
                        # Ignorer les lignes complètement vides
                        if any(c is not None for c in row):
                            data.append(row)
            
            wb.close()
            
            if headers and data:
                df = pd.DataFrame(data, columns=headers)
                # Convertir tout en string
                df = df.astype(str)
                df = df.replace('None', pd.NA)
        except Exception as e:
            df = None
        
        # Méthode 2: pandas standard si la méthode 1 échoue
        if df is None or len(df) == 0:
            try:
                df_test = pd.read_excel(io_module.BytesIO(file_content), dtype=str, nrows=2, engine='openpyxl')
                if df_test.columns[0].startswith('Unnamed'):
                    df = pd.read_excel(io_module.BytesIO(file_content), dtype=str, skiprows=1, engine='openpyxl')
                else:
                    df = pd.read_excel(io_module.BytesIO(file_content), dtype=str, engine='openpyxl')
            except:
                df = pd.read_excel(io_module.BytesIO(file_content), dtype=str)
    else:
        # CSV
        text = file_content.decode('utf-8', errors='ignore')
        try:
            df = pd.read_csv(io_module.StringIO(text), sep=None, engine='python', dtype=str, on_bad_lines='skip')
        except:
            df = pd.read_csv(io_module.StringIO(text), sep=',', dtype=str, on_bad_lines='skip')
    
    # === NORMALISER LES COLONNES ===
    # Supprimer les colonnes vides (Col_XX)
    df = df.loc[:, ~df.columns.str.match(r'^Col_\d+$')]
    
    # === FORMAT EPOD_TASK_LIST_V2 (nouvel export Cainiao) ===
    # Mapper les colonnes V2 vers les noms attendus par le reste de l'app
    v2_map = {
        "Waybill Number": "Tracking No.",
        "Zip Code": "Receiver's Zip Code",
        "The destination city": "Receiver's City",
        "Detailed address": "Receiver's Detail Address",
    }
    for src_col, dst_col in v2_map.items():
        if src_col in df.columns and dst_col not in df.columns:
            df[dst_col] = df[src_col]

    # Normaliser "Receiver's Zip Code" -> "Sort Code" si absent
    if 'Sort Code' not in df.columns and "Receiver's Zip Code" in df.columns:
        df['Sort Code'] = df["Receiver's Zip Code"]
    
    # Nettoyer les codes postaux (enlever apostrophes et ajouter 0 manquant)
    if 'Sort Code' in df.columns:
        df['Sort Code'] = df['Sort Code'].astype(str).str.strip().str.lstrip("'").str.strip()
        # Ajouter le 0 devant les codes postaux à 4 chiffres (ex: 2160 -> 02160)
        df['Sort Code'] = df['Sort Code'].apply(lambda x: '0' + x if pd.notna(x) and str(x) not in ['', 'nan', 'None', 'NA'] and str(x).isdigit() and len(str(x)) == 4 else x)
    
    # Parser GPS
    def split_gps(val):
        try:
            if pd.isna(val) or ',' not in str(val): return None, None
            lat, lon = str(val).replace('"', '').split(',')
            return float(lat), float(lon)
        except: return None, None
    
    gps_columns = ["Receiver to (Latitude,Longitude)", "GPS", "Coordinates", "LatLng"]
    has_gps_column = False
    # Format V2 : latitude et longitude dans deux colonnes séparées
    if "Receiver to Latitude" in df.columns and "Receiver to Longitude" in df.columns:
        df['lat'] = df["Receiver to Latitude"]
        df['lon'] = df["Receiver to Longitude"]
        has_gps_column = True
    else:
        for col in gps_columns:
            if col in df.columns:
                df[['lat', 'lon']] = df[col].apply(lambda x: pd.Series(split_gps(x)))
                has_gps_column = True
                break
    
    if 'lat' in df.columns:
        df['lat'] = pd.to_numeric(df['lat'], errors='coerce')
    if 'lon' in df.columns:
        df['lon'] = pd.to_numeric(df['lon'], errors='coerce')
    
    # === GÉOCODAGE PAR CODE POSTAL si pas de GPS ===
    if not has_gps_column or ('lat' in df.columns and df['lat'].isna().all()):
        df = geocode_by_postal_code(df)
    elif 'lat' in df.columns and df['lat'].isna().any():
        # Géocoder seulement les colis sans GPS
        mask_no_gps = df['lat'].isna()
        if mask_no_gps.any():
            df_no_gps = geocode_by_postal_code(df[mask_no_gps].copy())
            df.loc[mask_no_gps, 'lat'] = df_no_gps['lat']
            df.loc[mask_no_gps, 'lon'] = df_no_gps['lon']
    
    # === GÉOCODAGE INVERSE pour adresses censurées (******) ===
    addr_col = None
    for col in ["Receiver's Detail Address", "Receivers Detail Address", "Address"]:
        if col in df.columns:
            addr_col = col
            break
    
    if addr_col and 'lat' in df.columns and 'lon' in df.columns:
        # Détecter les adresses censurées (contiennent * ou sont vides)
        mask_censored = df[addr_col].apply(lambda x: '*' in str(x) if pd.notna(x) else True)
        mask_has_gps = df['lat'].notna() & df['lon'].notna()
        mask_to_reverse = mask_censored & mask_has_gps
        
        if mask_to_reverse.any():
            df = reverse_geocode_addresses(df, addr_col, mask_to_reverse)
    
    return df


def reverse_geocode_addresses(df, addr_col, mask):
    """Récupère les adresses réelles à partir des coordonnées GPS."""
    import urllib.request
    import urllib.parse
    
    # Cache pour éviter les appels dupliqués (même coordonnées)
    reverse_cache = {}
    
    # Collecter les coordonnées uniques
    coords_to_lookup = []
    for idx in df[mask].index:
        lat = df.at[idx, 'lat']
        lon = df.at[idx, 'lon']
        if pd.notna(lat) and pd.notna(lon):
            coords_to_lookup.append((idx, round(float(lat), 6), round(float(lon), 6)))
    
    # Géocoder par batch (limiter les appels API)
    for idx, lat, lon in coords_to_lookup:
        cache_key = f"{lat}_{lon}"
        
        if cache_key not in reverse_cache:
            try:
                params = urllib.parse.urlencode({
                    'lat': lat,
                    'lon': lon
                })
                url = f"https://api-adresse.data.gouv.fr/reverse/?{params}"
                
                req = urllib.request.Request(url, headers={'User-Agent': 'JNR-Dispatch/1.0'})
                with urllib.request.urlopen(req, timeout=5) as response:
                    result = json.loads(response.read().decode())
                    
                    if result.get('features'):
                        props = result['features'][0]['properties']
                        reverse_cache[cache_key] = {
                            'address': props.get('name', ''),
                            'city': props.get('city', ''),
                            'postcode': props.get('postcode', ''),
                            'label': props.get('label', '')
                        }
                    else:
                        reverse_cache[cache_key] = None
            except:
                reverse_cache[cache_key] = None
        
        # Appliquer l'adresse trouvée
        if reverse_cache.get(cache_key):
            addr_info = reverse_cache[cache_key]
            df.at[idx, addr_col] = addr_info['address']
            
            # Mettre à jour la ville si elle est aussi censurée
            for city_col in ["Receiver's City", "Receivers City"]:
                if city_col in df.columns:
                    current_city = str(df.at[idx, city_col])
                    if '*' in current_city or pd.isna(df.at[idx, city_col]):
                        df.at[idx, city_col] = addr_info['city']
                    break
    
    return df


def geocode_by_postal_code(df):
    """Géocode les colis par code postal + ville."""
    if 'Sort Code' not in df.columns:
        return df
    
    if 'lat' not in df.columns:
        df['lat'] = pd.NA
    if 'lon' not in df.columns:
        df['lon'] = pd.NA
    
    # Trouver la colonne ville
    city_col = None
    for col in ["Receiver's City", "Receivers City", "City", "Receiver's Region/Province"]:
        if col in df.columns:
            city_col = col
            break
    
    # Construire les requêtes uniques (CP + Ville)
    geocode_cache = {}
    
    unique_locations = set()
    for _, row in df.iterrows():
        cp = str(row.get('Sort Code', '')).strip()
        city = str(row.get(city_col, '')).strip() if city_col else ''
        if cp and cp != 'nan':
            unique_locations.add((cp, city))
    
    # Géocoder via l'API BAN (Base Adresse Nationale)
    import urllib.request
    import urllib.parse
    
    for cp, city in unique_locations:
        cache_key = f"{cp}_{city}"
        if cache_key in geocode_cache:
            continue
        
        try:
            query = f"{city}" if city and city != 'nan' else cp
            params = urllib.parse.urlencode({
                'q': query,
                'postcode': cp,
                'limit': 1
            })
            url = f"https://api-adresse.data.gouv.fr/search/?{params}"
            
            req = urllib.request.Request(url, headers={'User-Agent': 'JNR-Dispatch/1.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                result = json.loads(response.read().decode())
                
                if result.get('features'):
                    coords = result['features'][0]['geometry']['coordinates']
                    geocode_cache[cache_key] = (coords[1], coords[0])
                else:
                    geocode_cache[cache_key] = (None, None)
        except:
            geocode_cache[cache_key] = (None, None)
    
    # Appliquer les coordonnées
    for idx, row in df.iterrows():
        if pd.notna(row.get('lat')) and pd.notna(row.get('lon')):
            continue
        
        cp = str(row.get('Sort Code', '')).strip()
        city = str(row.get(city_col, '')).strip() if city_col else ''
        cache_key = f"{cp}_{city}"
        
        if cache_key in geocode_cache:
            lat, lon = geocode_cache[cache_key]
            if lat is not None:
                df.at[idx, 'lat'] = lat
                df.at[idx, 'lon'] = lon
    
    # Vérifier combien ont été géocodés
    geocoded = df['lat'].notna().sum()
    total = len(df)
    if geocoded < total:
        # Fallback: utiliser les coordonnées du centre du code postal
        cp_centers = {}
        for idx, row in df.iterrows():
            if pd.notna(row.get('lat')):
                cp = str(row.get('Sort Code', '')).strip()
                if cp not in cp_centers:
                    cp_centers[cp] = []
                cp_centers[cp].append((float(row['lat']), float(row['lon'])))
        
        # Calculer les centres
        for cp, coords in cp_centers.items():
            avg_lat = sum(c[0] for c in coords) / len(coords)
            avg_lon = sum(c[1] for c in coords) / len(coords)
            cp_centers[cp] = (avg_lat, avg_lon)
        
        # Appliquer aux colis restants
        for idx, row in df.iterrows():
            if pd.isna(row.get('lat')):
                cp = str(row.get('Sort Code', '')).strip()
                if cp in cp_centers:
                    df.at[idx, 'lat'] = cp_centers[cp][0]
                    df.at[idx, 'lon'] = cp_centers[cp][1]
    
    return df

# === FONCTIONS UTILITAIRES ===

def load_patterns():
    """Charge les patterns sauvegardés depuis le fichier JSON."""
    if os.path.exists(PATTERNS_FILE):
        with open(PATTERNS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"drivers": {}, "updated_at": None}

def save_patterns(patterns):
    """Sauvegarde les patterns dans le fichier JSON."""
    patterns["updated_at"] = datetime.now().isoformat()
    with open(PATTERNS_FILE, 'w', encoding='utf-8') as f:
        json.dump(patterns, f, ensure_ascii=False, indent=2)

def get_driver_color(index):
    """Retourne une couleur unique pour chaque chauffeur."""
    colors = [
        "#e74c3c", "#3498db", "#2ecc71", "#f39c12", "#9b59b6",
        "#1abc9c", "#e67e22", "#34495e", "#16a085", "#c0392b",
        "#2980b9", "#27ae60", "#d35400", "#8e44ad", "#17a2b8"
    ]
    return colors[index % len(colors)]

def normalize_text(text):
    """Normalise le texte pour comparaison (accents, casse, tirets, espaces)."""
    if not text or pd.isna(text):
        return ""
    text = str(text).lower().strip()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[-_\s]+', ' ', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    return text.strip()

def levenshtein_distance(s1, s2):
    """Calcule la distance de Levenshtein entre deux chaînes."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    
    return previous_row[-1]

def fuzzy_match_city(city_input, city_list, max_distance=2):
    """Vérifie si une ville correspond à la liste avec tolérance aux fautes."""
    if not city_list:
        return False
    if city_input is None or pd.isna(city_input) or str(city_input).strip() == '':
        return False
    
    normalized_input = normalize_text(city_input)
    if not normalized_input:
        return False
    
    for city in city_list:
        normalized_city = normalize_text(city)
        if not normalized_city:
            continue
        
        if normalized_input == normalized_city:
            return True
        
        if normalized_input in normalized_city or normalized_city in normalized_input:
            return True
        
        tolerance = min(max_distance, max(1, len(normalized_city) // 4))
        if levenshtein_distance(normalized_input, normalized_city) <= tolerance:
            return True
    
    return False

def match_postal_code(sort_code, postal_codes):
    """Vérifie si un code postal correspond à la liste assignée."""
    if not postal_codes:
        return False
    if sort_code is None or pd.isna(sort_code) or str(sort_code).strip() == '':
        return False
    
    # Nettoyer le code postal (enlever apostrophes, espaces, leading zeros)
    sort_code_str = str(sort_code).strip().lstrip("'").strip()
    sort_code_clean = sort_code_str.lstrip('0') if sort_code_str.startswith('0') else sort_code_str
    
    for cp in postal_codes:
        cp_str = str(cp).strip().lstrip("'").strip()
        cp_clean = cp_str.lstrip('0') if cp_str.startswith('0') else cp_str
        
        # Match exact
        if sort_code_str == cp_str or sort_code_clean == cp_clean:
            return True
        
        # Match par préfixe (ex: "51" matche "51100", "51200", etc.)
        if len(cp_str) < 5 and (sort_code_str.startswith(cp_str) or sort_code_clean.startswith(cp_clean)):
            return True
    
    return False

def point_in_zones(lat, lon, zones):
    """Vérifie si un point est dans une des zones géographiques."""
    if not zones:
        return False
    point = Point(lon, lat)
    for zone in zones:
        try:
            polygon = shape(zone)
            if polygon.contains(point):
                return True
        except:
            continue
    return False

def match_driver(row, driver_data):
    """Vérifie si un colis correspond aux critères d'un chauffeur."""
    # 1. Vérifier les codes postaux
    postal_codes = driver_data.get("postal_codes", [])
    sort_code = row.get('Sort Code', '')
    if match_postal_code(sort_code, postal_codes):
        return True
    
    # 2. Vérifier les villes
    cities = driver_data.get("cities", [])
    city_columns = ["Receiver's City", "Receivers City", "City", "Ville", "Receiver's Region/Province"]
    for col in city_columns:
        if col in row.index:
            city_value = row.get(col, '')
            if fuzzy_match_city(city_value, cities):
                return True
    
    # 3. Vérifier les zones géographiques (si coordonnées disponibles)
    zones = driver_data.get("zones", [])
    lat = row.get('lat')
    lon = row.get('lon')
    if pd.notna(lat) and pd.notna(lon) and zones:
        if point_in_zones(float(lat), float(lon), zones):
            return True
    
    return False

def auto_dispatch(df, patterns):
    """Dispatch automatique basé sur les patterns sauvegardés."""
    results = {}
    assigned_indices = set()
    
    for driver_name, driver_data in patterns.get("drivers", {}).items():
        has_criteria = (
            driver_data.get("zones", []) or 
            driver_data.get("postal_codes", []) or 
            driver_data.get("cities", [])
        )
        if not has_criteria:
            continue
        
        mask = df.apply(lambda row: match_driver(row, driver_data), axis=1)
        driver_df = df[mask]
        driver_df = driver_df[~driver_df.index.isin(assigned_indices)]
        
        if not driver_df.empty:
            results[driver_name] = driver_df
            assigned_indices.update(driver_df.index.tolist())
    
    unassigned = df[~df.index.isin(assigned_indices)]
    if not unassigned.empty:
        results["_NON_ASSIGNES"] = unassigned
    
    return results

def create_zip_with_excels(dispatch_results):
    """Crée un ZIP contenant tous les fichiers Excel."""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for driver_name, driver_df in dispatch_results.items():
            if driver_df.empty:
                continue
            excel_data = preparer_telechargement_excel(driver_df)
            safe_name = driver_name.replace(" ", "_").replace("/", "-")
            filename = f"{safe_name}.xlsx"
            zip_file.writestr(filename, excel_data)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def get_driver_summary(driver_data):
    """Génère un résumé des critères d'un chauffeur."""
    parts = []
    
    zones = driver_data.get("zones", [])
    if zones:
        parts.append(f"{len(zones)} zone(s)")
    
    postal_codes = driver_data.get("postal_codes", [])
    if postal_codes:
        parts.append(f"CP: {', '.join(postal_codes[:3])}{'...' if len(postal_codes) > 3 else ''}")
    
    cities = driver_data.get("cities", [])
    if cities:
        parts.append(f"Villes: {', '.join(cities[:2])}{'...' if len(cities) > 2 else ''}")
    
    return " | ".join(parts) if parts else "Aucun critère"

# === INTERFACE ===

st.title("🚚 Dispatch Automatique - JNR Transport")

patterns = load_patterns()

tab1, tab2, tab3 = st.tabs(["📍 Zones Géographiques", "🏘️ Codes Postaux & Villes", "⚡ Dispatch Automatique"])

# === TAB 1: CONFIGURATION DES ZONES GÉOGRAPHIQUES ===
with tab1:
    st.markdown("### Définir les zones de livraison sur la carte")
    
    # Sous-tabs pour les deux modes
    subtab1, subtab2 = st.tabs(["✏️ Dessiner des zones", "🔧 Gérer les zones"])
    
    # === SOUS-TAB 1: DESSINER ===
    with subtab1:
        col_options = st.columns([2, 1, 1])
        with col_options[0]:
            uploaded_ref = st.file_uploader(
                "Charger un fichier de référence", 
                type=['csv', 'xlsx', 'xls'],
                key="ref_file"
            )
        with col_options[1]:
            sample_rate = st.selectbox(
                "Afficher 1 point sur",
                options=[1, 5, 10, 20],
                index=2,
                help="Réduire pour plus de fluidité"
            )
        with col_options[2]:
            show_points = st.checkbox("Afficher les points", value=True)
        
        # Charger les données pour avoir la liste des CP
        df_map = pd.DataFrame()
        available_cp = []
        if uploaded_ref:
            file_content = uploaded_ref.getvalue()
            df_ref = load_and_process_file(file_content, uploaded_ref.name)
            if 'lat' in df_ref.columns and 'lon' in df_ref.columns:
                df_map = df_ref.dropna(subset=['lat', 'lon']).copy()
            if 'Sort Code' in df_ref.columns:
                available_cp = sorted(df_ref['Sort Code'].dropna().unique().tolist())
        
        # Filtre par codes postaux
        selected_cp = []
        if available_cp:
            with st.expander("🔍 Filtrer par codes postaux", expanded=False):
                col_filter1, col_filter2 = st.columns([3, 1])
                with col_filter1:
                    selected_cp = st.multiselect(
                        "Afficher seulement ces codes postaux:",
                        options=available_cp,
                        default=[],
                        placeholder="Tous les CP (cliquer pour filtrer)",
                        key="cp_filter"
                    )
                with col_filter2:
                    if selected_cp:
                        st.metric("CP sélectionnés", len(selected_cp))
                    else:
                        st.metric("CP affichés", len(available_cp))
            
            # Filtrer le dataframe si des CP sont sélectionnés
            if selected_cp and not df_map.empty:
                df_map = df_map[df_map['Sort Code'].isin(selected_cp)]
        
        col_left, col_right = st.columns([3, 1])
        
        with col_right:
            st.markdown("#### 👥 Chauffeurs")
            
            new_driver = st.text_input("Nom du chauffeur", placeholder="Ex: Mohamed", key="new_driver_tab1")
            if st.button("➕ Ajouter", use_container_width=True, key="add_driver_tab1"):
                if new_driver and new_driver.strip():
                    driver_name = new_driver.strip()
                    if driver_name not in patterns.get("drivers", {}):
                        if "drivers" not in patterns:
                            patterns["drivers"] = {}
                        patterns["drivers"][driver_name] = {
                            "zones": [], 
                            "postal_codes": [],
                            "cities": [],
                            "color": get_driver_color(len(patterns["drivers"]))
                        }
                        save_patterns(patterns)
                        st.success(f"✅ {driver_name} ajouté!")
                        st.rerun()
                    else:
                        st.warning("Ce chauffeur existe déjà")
            
            st.markdown("---")
            
            selected_driver = st.selectbox(
                "Chauffeur à configurer:",
                options=list(patterns.get("drivers", {}).keys()) or ["Aucun chauffeur"],
                key="driver_select_tab1"
            )
            
            st.markdown("#### 📊 Résumé")
            for driver, data in patterns.get("drivers", {}).items():
                color = data.get("color", "#666")
                summary = get_driver_summary(data)
                st.markdown(f"""
                    <div style="margin:4px 0; padding:8px; background:#f8f9fa; border-radius:4px; border-left:4px solid {color};">
                        <strong>{driver}</strong><br/>
                        <small style="color:#666;">{summary}</small>
                    </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            if selected_driver and selected_driver != "Aucun chauffeur":
                st.markdown(f"**Actions pour {selected_driver}:**")
                
                zones_count = len(patterns["drivers"].get(selected_driver, {}).get("zones", []))
                if zones_count > 0:
                    if st.button(f"🗑️ Supprimer les {zones_count} zone(s)", use_container_width=True):
                        patterns["drivers"][selected_driver]["zones"] = []
                        save_patterns(patterns)
                        st.success("Zones supprimées!")
                        st.rerun()
                
                if st.button("❌ Supprimer le chauffeur", use_container_width=True, key="del_driver_tab1"):
                    del patterns["drivers"][selected_driver]
                    save_patterns(patterns)
                    st.success("Chauffeur supprimé!")
                    st.rerun()
        
        with col_left:
            center_lat, center_lon = 49.25, 4.03
            
            if not df_map.empty:
                center_lat = df_map['lat'].mean()
                center_lon = df_map['lon'].mean()
            
            m = folium.Map(location=[center_lat, center_lon], zoom_start=10, prefer_canvas=True)
            
            Draw(
                export=False,
                draw_options={
                    'polyline': False, 
                    'circle': False, 
                    'marker': False, 
                    'circlemarker': False, 
                    'polygon': True, 
                    'rectangle': True
                }
            ).add_to(m)
            
            # Afficher les zones existantes
            for driver, data in patterns.get("drivers", {}).items():
                color = data.get("color", "#666")
                for zone in data.get("zones", []):
                    folium.GeoJson(
                        zone,
                        style_function=lambda x, c=color: {
                            'fillColor': c,
                            'color': c,
                            'weight': 2,
                            'fillOpacity': 0.3
                        },
                        tooltip=driver
                    ).add_to(m)
            
            # Afficher les points
            if show_points and not df_map.empty:
                df_sampled = df_map.iloc[::sample_rate]
                
                if sample_rate == 1:
                    # Mode 1/1 : afficher tous les points individuellement (sans clustering)
                    for _, row in df_sampled.iterrows():
                        folium.CircleMarker(
                            location=[row['lat'], row['lon']],
                            radius=5,
                            color='#333',
                            fill=True,
                            fillColor='#333',
                            fillOpacity=0.7,
                            weight=1
                        ).add_to(m)
                    st.caption(f"📍 {len(df_map)} points affichés")
                else:
                    # Mode échantillonné : utiliser le clustering pour la performance
                    callback = """
                    function (row) {
                        var marker = L.circleMarker(new L.LatLng(row[0], row[1]), {
                            radius: 4,
                            color: '#333',
                            fillColor: '#333',
                            fillOpacity: 0.6
                        });
                        return marker;
                    }
                    """
                    points_data = df_sampled[['lat', 'lon']].values.tolist()
                    FastMarkerCluster(data=points_data, callback=callback).add_to(m)
                    st.caption(f"📍 {len(df_sampled)}/{len(df_map)} points affichés (1 sur {sample_rate})")
            
            output = st_folium(m, width="100%", height=500, key="config_map", returned_objects=["all_drawings"])
            
            if output and output.get('all_drawings'):
                last_draw = output['all_drawings'][-1]
                if last_draw and 'geometry' in last_draw:
                    st.info(f"🎯 Zone détectée! Cliquez pour l'assigner à **{selected_driver}**")
                    
                    if st.button(f"✅ Assigner cette zone à {selected_driver}", type="primary"):
                        if selected_driver and selected_driver != "Aucun chauffeur":
                            geometry = last_draw['geometry']
                            if "zones" not in patterns["drivers"][selected_driver]:
                                patterns["drivers"][selected_driver]["zones"] = []
                            patterns["drivers"][selected_driver]["zones"].append(geometry)
                            save_patterns(patterns)
                            st.success(f"Zone ajoutée pour {selected_driver}!")
                            st.rerun()
    
    # === SOUS-TAB 2: GÉRER LES ZONES ===
    with subtab2:
        st.markdown("#### Réassigner ou supprimer des zones")
        
        total_zones = sum(len(d.get("zones", [])) for d in patterns.get("drivers", {}).values())
        
        if total_zones == 0:
            st.info("Aucune zone définie. Allez dans l'onglet 'Dessiner des zones' pour en créer.")
        else:
            col_manage_left, col_manage_right = st.columns([3, 1])
            
            with col_manage_right:
                st.markdown("#### 🎯 Sélection")
                
                drivers_with_zones = {d: data for d, data in patterns.get("drivers", {}).items() if data.get("zones")}
                
                if not drivers_with_zones:
                    st.warning("Aucune zone à gérer")
                else:
                    selected_manage_driver = st.selectbox(
                        "Chauffeur:",
                        options=list(drivers_with_zones.keys()),
                        key="manage_driver_select"
                    )
                    
                    if selected_manage_driver:
                        zones = patterns["drivers"][selected_manage_driver].get("zones", [])
                        selected_zone_idx = st.selectbox(
                            f"Zone de {selected_manage_driver}:",
                            options=range(len(zones)),
                            format_func=lambda x: f"Zone {x+1}",
                            key="manage_zone_select"
                        )
                        
                        st.markdown("---")
                        st.markdown("#### ⚙️ Actions")
                        
                        other_drivers = [d for d in patterns.get("drivers", {}).keys() if d != selected_manage_driver]
                        if other_drivers:
                            reassign_to = st.selectbox(
                                "Réassigner à:",
                                options=other_drivers,
                                key="reassign_to"
                            )
                            
                            if st.button(f"↔️ Réassigner à {reassign_to}", use_container_width=True):
                                zone_to_move = patterns["drivers"][selected_manage_driver]["zones"][selected_zone_idx]
                                patterns["drivers"][selected_manage_driver]["zones"].pop(selected_zone_idx)
                                if "zones" not in patterns["drivers"][reassign_to]:
                                    patterns["drivers"][reassign_to]["zones"] = []
                                patterns["drivers"][reassign_to]["zones"].append(zone_to_move)
                                save_patterns(patterns)
                                st.success(f"Zone réassignée à {reassign_to}!")
                                st.rerun()
                        
                        st.markdown("---")
                        
                        if st.button("🗑️ Supprimer cette zone", use_container_width=True, type="secondary"):
                            patterns["drivers"][selected_manage_driver]["zones"].pop(selected_zone_idx)
                            save_patterns(patterns)
                            st.success("Zone supprimée!")
                            st.rerun()
            
            with col_manage_left:
                center_lat, center_lon = 49.25, 4.03
                m_manage = folium.Map(location=[center_lat, center_lon], zoom_start=10, prefer_canvas=True)
                
                for driver, data in patterns.get("drivers", {}).items():
                    color = data.get("color", "#666")
                    zones = data.get("zones", [])
                    for idx, zone in enumerate(zones):
                        folium.GeoJson(
                            zone,
                            style_function=lambda x, c=color: {
                                'fillColor': c,
                                'color': c,
                                'weight': 2,
                                'fillOpacity': 0.3
                            },
                            tooltip=f"{driver} - Zone {idx+1}"
                        ).add_to(m_manage)
                
                st_folium(m_manage, width="100%", height=500, key="manage_map")

# === TAB 2: CODES POSTAUX & VILLES ===
with tab2:
    st.markdown("### Assigner des codes postaux et villes aux chauffeurs")
    st.caption("💡 Pour les chauffeurs qui couvrent des zones entières sans besoin de tracer sur la carte")
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.markdown("#### 👥 Chauffeurs")
        
        new_driver2 = st.text_input("Nom du chauffeur", placeholder="Ex: Mohamed", key="new_driver_tab2")
        if st.button("➕ Ajouter", use_container_width=True, key="add_driver_tab2"):
            if new_driver2 and new_driver2.strip():
                driver_name = new_driver2.strip()
                if driver_name not in patterns.get("drivers", {}):
                    if "drivers" not in patterns:
                        patterns["drivers"] = {}
                    patterns["drivers"][driver_name] = {
                        "zones": [], 
                        "postal_codes": [],
                        "cities": [],
                        "color": get_driver_color(len(patterns["drivers"]))
                    }
                    save_patterns(patterns)
                    st.success(f"✅ {driver_name} ajouté!")
                    st.rerun()
        
        st.markdown("---")
        
        selected_driver2 = st.selectbox(
            "Chauffeur à configurer:",
            options=list(patterns.get("drivers", {}).keys()) or ["Aucun chauffeur"],
            key="driver_select_tab2"
        )
    
    with col2:
        if selected_driver2 and selected_driver2 != "Aucun chauffeur":
            driver_data = patterns["drivers"].get(selected_driver2, {})
            color = driver_data.get("color", "#666")
            
            st.markdown(f"#### Configuration de <span style='color:{color}'>{selected_driver2}</span>", unsafe_allow_html=True)
            
            # === CODES POSTAUX ===
            st.markdown("##### 📮 Codes Postaux")
            current_cp = driver_data.get("postal_codes", [])
            
            cp_input = st.text_input(
                "Ajouter des codes postaux (séparés par virgules)",
                placeholder="Ex: 51100, 51110, 08",
                key=f"cp_input_{selected_driver2}"
            )
            
            col_cp1, col_cp2 = st.columns([3, 1])
            with col_cp1:
                if current_cp:
                    st.write("Actuels: " + ", ".join([f"`{cp}`" for cp in current_cp]))
                else:
                    st.caption("Aucun code postal assigné")
            
            with col_cp2:
                if st.button("➕ Ajouter CP", key=f"add_cp_{selected_driver2}"):
                    if cp_input:
                        new_cps = [cp.strip().lstrip("'") for cp in cp_input.split(",") if cp.strip()]
                        if "postal_codes" not in patterns["drivers"][selected_driver2]:
                            patterns["drivers"][selected_driver2]["postal_codes"] = []
                        for cp in new_cps:
                            if cp not in patterns["drivers"][selected_driver2]["postal_codes"]:
                                patterns["drivers"][selected_driver2]["postal_codes"].append(cp)
                        save_patterns(patterns)
                        st.success(f"Codes postaux ajoutés!")
                        st.rerun()
            
            if current_cp:
                if st.button("🗑️ Effacer tous les CP", key=f"clear_cp_{selected_driver2}"):
                    patterns["drivers"][selected_driver2]["postal_codes"] = []
                    save_patterns(patterns)
                    st.rerun()
            
            st.markdown("---")
            
            # === VILLES ===
            st.markdown("##### 🏘️ Villes & Villages")
            st.caption("Tolérant aux fautes de frappe et accents (Reims = reims = REIMS)")
            
            current_cities = driver_data.get("cities", [])
            
            cities_input = st.text_input(
                "Ajouter des villes (séparées par virgules)",
                placeholder="Ex: Reims, Épernay, Châlons-en-Champagne",
                key=f"cities_input_{selected_driver2}"
            )
            
            col_city1, col_city2 = st.columns([3, 1])
            with col_city1:
                if current_cities:
                    st.write("Actuelles: " + ", ".join([f"`{c}`" for c in current_cities]))
                else:
                    st.caption("Aucune ville assignée")
            
            with col_city2:
                if st.button("➕ Ajouter Villes", key=f"add_cities_{selected_driver2}"):
                    if cities_input:
                        new_cities = [c.strip() for c in cities_input.split(",") if c.strip()]
                        if "cities" not in patterns["drivers"][selected_driver2]:
                            patterns["drivers"][selected_driver2]["cities"] = []
                        for city in new_cities:
                            if city not in patterns["drivers"][selected_driver2]["cities"]:
                                patterns["drivers"][selected_driver2]["cities"].append(city)
                        save_patterns(patterns)
                        st.success(f"Villes ajoutées!")
                        st.rerun()
            
            if current_cities:
                if st.button("🗑️ Effacer toutes les villes", key=f"clear_cities_{selected_driver2}"):
                    patterns["drivers"][selected_driver2]["cities"] = []
                    save_patterns(patterns)
                    st.rerun()
            
            st.markdown("---")
            
            st.markdown("##### 📋 Résumé")
            zones_count = len(driver_data.get("zones", []))
            st.info(f"""
            **{selected_driver2}** recevra les colis qui correspondent à:
            - **{len(current_cp)}** code(s) postal(aux)
            - **{len(current_cities)}** ville(s)
            - **{zones_count}** zone(s) géographique(s)
            """)

# === TAB 3: DISPATCH AUTOMATIQUE ===
with tab3:
    st.markdown("### Importer et dispatcher automatiquement")
    
    total_criteria = 0
    for d in patterns.get("drivers", {}).values():
        total_criteria += len(d.get("zones", []))
        total_criteria += len(d.get("postal_codes", []))
        total_criteria += len(d.get("cities", []))
    
    if total_criteria == 0:
        st.warning("⚠️ Aucun critère n'est configuré. Configurez des zones, codes postaux ou villes pour vos chauffeurs.")
    else:
        st.success(f"✅ {len(patterns.get('drivers', {}))} chauffeur(s) configuré(s)")
        
        with st.expander("📋 Voir la configuration"):
            for driver, data in patterns.get("drivers", {}).items():
                summary = get_driver_summary(data)
                st.write(f"**{driver}**: {summary}")
    
    st.markdown("---")
    
    uploaded_dispatch = st.file_uploader(
        "📁 Charger le fichier Cainiao à dispatcher",
        type=['csv', 'xlsx', 'xls'],
        key="dispatch_file"
    )
    
    if uploaded_dispatch and total_criteria > 0:
        file_content = uploaded_dispatch.getvalue()
        df_dispatch = load_and_process_file(file_content, uploaded_dispatch.name)
        
        has_gps = 'lat' in df_dispatch.columns and df_dispatch['lat'].notna().any()
        has_city = "Receiver's City" in df_dispatch.columns or "Receivers City" in df_dispatch.columns
        has_cp = "Sort Code" in df_dispatch.columns
        
        st.info(f"""
        📦 **{len(df_dispatch)}** colis chargés
        - GPS: {'✅' if has_gps else '❌'}
        - Ville: {'✅' if has_city else '❌'}  
        - Code Postal: {'✅' if has_cp else '❌'}
        """)
        
        if st.button("🚀 Lancer le dispatch automatique", type="primary", use_container_width=True):
            with st.spinner("Dispatch en cours..."):
                results = auto_dispatch(df_dispatch, patterns)
            
            st.markdown("### 📊 Résultats du dispatch")
            
            cols = st.columns(3)
            col_idx = 0
            
            total_assigned = 0
            for driver_name, driver_df in results.items():
                if driver_name == "_NON_ASSIGNES":
                    continue
                
                with cols[col_idx % 3]:
                    color = patterns["drivers"].get(driver_name, {}).get("color", "#666")
                    st.markdown(f"""
                        <div style="padding:15px; background:white; border-radius:8px; border-left:5px solid {color}; margin-bottom:10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                            <h4 style="margin:0; color:{color};">{driver_name}</h4>
                            <p style="font-size:24px; font-weight:bold; margin:5px 0;">{len(driver_df)} colis</p>
                        </div>
                    """, unsafe_allow_html=True)
                    total_assigned += len(driver_df)
                col_idx += 1
            
            if "_NON_ASSIGNES" in results:
                unassigned = results["_NON_ASSIGNES"]
                st.warning(f"⚠️ **{len(unassigned)}** colis non assignés")
                
                with st.expander("Voir les colis non assignés"):
                    display_cols = [c for c in ["Tracking No.", "Sort Code", "Receiver's City", "Receiver's Detail Address"] if c in unassigned.columns]
                    if display_cols:
                        st.dataframe(unassigned[display_cols].head(100))
                    else:
                        st.dataframe(unassigned.head(100))
            
            st.markdown("---")
            st.markdown("### 📥 Télécharger les fichiers")
            
            zip_data = create_zip_with_excels(results)
            st.download_button(
                label="📦 Télécharger TOUS les fichiers (ZIP)",
                data=zip_data,
                file_name=f"Dispatch_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                mime="application/zip",
                use_container_width=True
            )
            
            st.markdown("---")
            st.markdown("**Ou télécharger individuellement:**")
            
            dl_cols = st.columns(3)
            dl_idx = 0
            for driver_name, driver_df in results.items():
                if driver_df.empty:
                    continue
                
                with dl_cols[dl_idx % 3]:
                    display_name = "Non assignés" if driver_name == "_NON_ASSIGNES" else driver_name
                    excel_data = preparer_telechargement_excel(driver_df)
                    
                    st.download_button(
                        label=f"📄 {display_name} ({len(driver_df)})",
                        data=excel_data,
                        file_name=f"{driver_name.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{driver_name}"
                    )
                dl_idx += 1

# === SIDEBAR ===
with st.sidebar:
    st.markdown("## ⚙️ Gestion")
    
    if patterns.get("drivers"):
        st.download_button(
            label="💾 Exporter la config",
            data=json.dumps(patterns, ensure_ascii=False, indent=2),
            file_name="driver_patterns_backup.json",
            mime="application/json",
            use_container_width=True
        )
    
    st.markdown("---")
    uploaded_patterns = st.file_uploader("📂 Importer une config", type=['json'], key="import_patterns")
    if uploaded_patterns:
        try:
            imported = json.load(uploaded_patterns)
            if st.button("✅ Appliquer cette configuration"):
                save_patterns(imported)
                st.success("Configuration importée!")
                st.rerun()
        except:
            st.error("Fichier JSON invalide")
    
    st.markdown("---")
    st.markdown("### 📖 Guide")
    st.markdown("""
    **3 façons d'assigner:**
    1. 📍 **Zones** - Dessiner sur la carte
    2. 📮 **Codes Postaux** - Ex: 51100, 08
    3. 🏘️ **Villes** - Tolérant aux fautes
    
    Les critères se cumulent!
    """)
    
    if patterns.get("updated_at"):
        st.caption(f"Mis à jour: {patterns['updated_at'][:16]}")
